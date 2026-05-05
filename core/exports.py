"""
نظام تصدير البيانات - Excel و PDF
يوفر وظائف تصدير حديثة للتقارير والكتب

متطلبات:
  pip install openpyxl==3.1.2
  
الحالة:
  ✅ مثبت في requirements.txt
  ✅ جاهز للاستخدام
"""

import io
from datetime import datetime
from django.http import HttpResponse
from django.utils import timezone

# ملاحظة: openpyxl مثبت (يمكنك تجاهل تحذير VS Code)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    raise ImportError(
        "مكتبة openpyxl غير مثبتة. الرجاء تشغيل:\n"
        "pip install -r requirements.txt"
    ) from e


def export_books_to_excel(books, filename="books_export.xlsx"):
    """
    تصدير قائمة الكتب إلى ملف Excel مع تنسيق احترافي
    """
    # إنشاء workbook جديد
    wb = Workbook()
    ws = wb.active
    ws.title = "الكتب"
    
    # تعريف الألوان
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    incoming_fill = PatternFill(start_color="D5E3F7", end_color="D5E3F7", fill_type="solid")
    outgoing_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # تعريف الحدود
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # العناوين
    headers = [
        'الرقم',
        'التاريخ',
        'النوع',
        'رقم الكتاب',
        'العنوان',
        'الجهة',
        'الحالة',
        'تاريخ الاستحقاق',
        'الوصف',
        'المستخدم',
    ]
    
    # كتابة العناوين
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # كتابة البيانات
    for row_idx, book in enumerate(books, start=2):
        # البيانات
        data = [
            row_idx - 1,
            book.date.strftime('%Y-%m-%d') if book.date else '',
            'وارد' if book.is_incoming else 'صادر',
            book.number or '',
            book.title or '',
            book.entity.name if book.entity else '',
            {'pending': 'قيد المتابعة', 'done': 'منجز', 'hold': 'معلق'}.get(book.final_status, ''),
            book.due_date.strftime('%Y-%m-%d') if book.due_date else '',
            book.description or '',
            book.created_by.get_full_name() or book.created_by.username if book.created_by else '',
        ]
        
        # تحديد لون الصف حسب نوع الكتاب
        row_fill = incoming_fill if book.is_incoming else outgoing_fill
        
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = thin_border
            cell.fill = row_fill
    
    # ضبط عرض الأعمدة
    column_widths = [8, 12, 10, 15, 30, 25, 15, 15, 40, 20]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # حفظ في buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # إنشاء response
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def export_report_to_excel(data, filename="report_export.xlsx"):
    """
    تصدير التقارير الإحصائية إلى Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "التقرير"
    
    # تنسيق العنوان الرئيسي
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # العنوان الرئيسي
    ws.merge_cells('A1:B1')
    title_cell = ws['A1']
    title_cell.value = f"تقرير نظام إدارة الكتب - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = header_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # كتابة البيانات
    current_row = 3
    for section, items in data.items():
        # عنوان القسم
        ws.merge_cells(f'A{current_row}:B{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = section
        section_cell.font = Font(bold=True, size=12)
        section_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        section_cell.alignment = Alignment(horizontal='center')
        section_cell.border = thin_border
        current_row += 1
        
        # عناصر القسم
        for key, value in items.items():
            ws[f'A{current_row}'] = key
            ws[f'B{current_row}'] = value
            ws[f'A{current_row}'].alignment = Alignment(horizontal='right')
            ws[f'B{current_row}'].alignment = Alignment(horizontal='center')
            ws[f'A{current_row}'].border = thin_border
            ws[f'B{current_row}'].border = thin_border
            current_row += 1
        
        current_row += 1  # فراغ بين الأقسام
    
    # ضبط عرض الأعمدة
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    
    # حفظ في buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def get_export_data(books):
    """
    تحضير البيانات للتصدير
    """
    data = []
    for book in books:
        data.append({
            'date': book.date,
            'kind': 'وارد' if book.kind == 'incoming' else 'صادر',
            'number': book.number or '',
            'title': book.title or '',
            'entity': book.entity.name if book.entity else '',
            'status': {
                'pending': 'قيد المتابعة',
                'done': 'منجز',
                'hold': 'معلق'
            }.get(book.final_status, ''),
            'due_date': book.due_date,
            'description': book.description or '',
            'created_by': book.created_by.get_full_name() or book.created_by.username if book.created_by else '',
        })
    
    return data
